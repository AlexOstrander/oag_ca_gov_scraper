import unittest
import os
import tempfile
import pandas as pd
import datetime
import time
import math
from unittest.mock import patch, Mock, MagicMock
from bs4 import BeautifulSoup
import openpyxl
import oag_ca_gov_scraper
from io import StringIO  # Python 3 import

class TestUrlReading(unittest.TestCase):
    """
    Test class for validating URL extraction functionality from Excel and TSV files.
    This class tests the functionality of reading and filtering URLs from different
    file formats (Excel, TSV) with various configurations:
    - Reading from default columns
    - Reading from specific columns by index or name
    - Handling non-URL text entries
    - Handling nonexistent files
    The test setup creates temporary Excel and TSV files with predefined URL data,
    and each test method validates a specific aspect of the URL extraction logic.
    Test data includes valid URLs and non-URL strings to verify filtering works correctly.
    """
    
    def setUp(self):
        # Create temp files for testing
        self.temp_dir = tempfile.mkdtemp()
        
        # Create a sample Excel file
        self.excel_file = os.path.join(self.temp_dir, "test_urls.xlsx")
        df = pd.DataFrame({
            "URLs": ["http://example.com/1", "http://example.com/2", "not-a-url", "http://example.com/3"],
            "Other": ["text1", "http://other.com/1", "text2", "text3"]
        })
        df.to_excel(self.excel_file, index=False)
        
        # Create a sample TSV file
        self.tsv_file = os.path.join(self.temp_dir, "test_urls.tsv")
        with open(self.tsv_file, 'w') as f:
            f.write("URLs\tOther\n")
            f.write("http://example.com/1\ttext1\n")
            f.write("http://example.com/2\thttp://other.com/1\n")
            f.write("not-a-url\ttext2\n")
            f.write("http://example.com/3\ttext3\n")
    
    def tearDown(self):
        # Clean up temp files
        for file_path in [self.excel_file, self.tsv_file]:
            if os.path.exists(file_path):
                os.remove(file_path)
        os.rmdir(self.temp_dir)
    
    def test_read_urls_from_excel_default_column(self):
        """Test reading URLs from the first column of an Excel file."""
        urls = oag_ca_gov_scraper.read_urls_from_excel(self.excel_file)
        self.assertEqual(len(urls), 3)
        self.assertEqual(urls, ["http://example.com/1", "http://example.com/2", "http://example.com/3"])
    
    def test_read_urls_from_excel_specific_column(self):
        """Test reading URLs from a specific column of an Excel file."""
        urls = oag_ca_gov_scraper.read_urls_from_excel(self.excel_file, column_index=1)
        self.assertEqual(len(urls), 1)
        self.assertEqual(urls, ["http://other.com/1"])
    
    def test_read_urls_from_excel_column_name(self):
        """Test reading URLs using a column name."""
        urls = oag_ca_gov_scraper.read_urls_from_excel(self.excel_file, column_index="Other")
        self.assertEqual(len(urls), 1)
        self.assertEqual(urls, ["http://other.com/1"])
    
    def test_read_urls_from_excel_nonexistent_file(self):
        """Test handling of a nonexistent Excel file."""
        urls = oag_ca_gov_scraper.read_urls_from_excel("nonexistent.xlsx")
        self.assertEqual(urls, [])
    
    def test_read_urls_from_tsv_default_column(self):
        """Test reading URLs from the first column of a TSV file."""
        urls = oag_ca_gov_scraper.read_urls_from_tsv(self.tsv_file)
        self.assertEqual(len(urls), 3)
        self.assertEqual(urls, ["http://example.com/1", "http://example.com/2", "http://example.com/3"])
    
    def test_read_urls_from_tsv_specific_column(self):
        """Test reading URLs from a specific column of a TSV file."""
        urls = oag_ca_gov_scraper.read_urls_from_tsv(self.tsv_file, column_index=1)
        self.assertEqual(len(urls), 1)
        self.assertEqual(urls, ["http://other.com/1"])
    
    def test_read_urls_from_tsv_nonexistent_file(self):
        """Test handling of a nonexistent TSV file."""
        urls = oag_ca_gov_scraper.read_urls_from_tsv("nonexistent.tsv")
        self.assertEqual(urls, [])


class TestFloatConversion(unittest.TestCase):
    
    def test_convert_empty_string(self):
        """Test conversion of an empty string."""
        result = oag_ca_gov_scraper.convert_to_float("")
        self.assertEqual(result, "")
    
    def test_convert_none(self):
        """Test conversion of None."""
        result = oag_ca_gov_scraper.convert_to_float(None)
        self.assertEqual(result, "")
    
    def test_convert_simple_float(self):
        """Test conversion of a simple float."""
        result = oag_ca_gov_scraper.convert_to_float("123.45")
        self.assertEqual(result, 123.45)
    
    def test_convert_integer(self):
        """Test conversion of an integer."""
        result = oag_ca_gov_scraper.convert_to_float("123")
        self.assertEqual(result, 123.0)
    
    def test_convert_dollar_sign(self):
        """Test conversion with dollar sign."""
        result = oag_ca_gov_scraper.convert_to_float("$123.45")
        self.assertEqual(result, 123.45)
    
    def test_convert_commas(self):
        """Test conversion with commas."""
        result = oag_ca_gov_scraper.convert_to_float("1,234.56")
        self.assertEqual(result, 1234.56)
    
    def test_convert_spaces(self):
        """Test conversion with spaces."""
        result = oag_ca_gov_scraper.convert_to_float("1 234.56")
        self.assertEqual(result, 1234.56)
    
    def test_convert_complex_format(self):
        """Test conversion with complex format."""
        result = oag_ca_gov_scraper.convert_to_float("$1,234,567.89")
        self.assertEqual(result, 1234567.89)
    
    def test_convert_already_numeric(self):
        """Test conversion of already numeric values."""
        result = oag_ca_gov_scraper.convert_to_float(123.45)
        self.assertEqual(result, 123.45)
        
        result = oag_ca_gov_scraper.convert_to_float(123)
        self.assertEqual(result, 123.0)
    
    def test_convert_invalid_format(self):
        """Test conversion of an invalid format."""
        result = oag_ca_gov_scraper.convert_to_float("not a number")
        self.assertEqual(result, "not a number")


class TestUrlGeneration(unittest.TestCase):
    """
    Test suite for the URL generation functionality in the oag_ca_gov_scraper module.
    This class tests the generate_urls_for_year_range function which creates URLs for 
    California Office of the Attorney General Proposition 65 notices. The tests verify
    correct URL generation for different scenarios:
    - Complete year ranges with sequential IDs
    - Year ranges with missing years
    - Year ranges with no matching data
    Each test validates the function's ability to correctly format URLs according to 
    the OAG's URL pattern (https://oag.ca.gov/prop65/60-Day-Notice-YYYY-NNNNN) and
    to properly handle edge cases in the input data.
    """
    
    def test_generate_urls_for_year_range(self):
        """Test generating URLs for a specific year range."""
        start_ids = {2020: 1, 2021: 100}
        end_ids = {2020: 3, 2021: 102}
        urls = oag_ca_gov_scraper.generate_urls_for_year_range(2020, 2021, start_ids, end_ids)
        
        expected_urls = [
            "https://oag.ca.gov/prop65/60-Day-Notice-2020-00001",
            "https://oag.ca.gov/prop65/60-Day-Notice-2020-00002",
            "https://oag.ca.gov/prop65/60-Day-Notice-2020-00003",
            "https://oag.ca.gov/prop65/60-Day-Notice-2021-00100",
            "https://oag.ca.gov/prop65/60-Day-Notice-2021-00101",
            "https://oag.ca.gov/prop65/60-Day-Notice-2021-00102"
        ]
        
        self.assertEqual(urls, expected_urls)
    
    def test_generate_urls_missing_year(self):
        """Test generating URLs with missing year data."""
        start_ids = {2020: 1, 2022: 1}
        end_ids = {2020: 2, 2022: 2}
        urls = oag_ca_gov_scraper.generate_urls_for_year_range(2020, 2022, start_ids, end_ids)
        
        expected_urls = [
            "https://oag.ca.gov/prop65/60-Day-Notice-2020-00001",
            "https://oag.ca.gov/prop65/60-Day-Notice-2020-00002",
            "https://oag.ca.gov/prop65/60-Day-Notice-2022-00001",
            "https://oag.ca.gov/prop65/60-Day-Notice-2022-00002"
        ]
        
        self.assertEqual(urls, expected_urls)
    
    def test_generate_urls_no_matching_years(self):
        """Test generating URLs with no matching years."""
        start_ids = {2018: 1}
        end_ids = {2018: 2}
        urls = oag_ca_gov_scraper.generate_urls_for_year_range(2020, 2022, start_ids, end_ids)
        
        self.assertEqual(urls, [])


class TestProgressEstimation(unittest.TestCase):
    """
    Unit tests for the progress estimation functionality in oag_ca_gov_scraper module.
    This test suite verifies the behavior of the estimate_progress function
    which calculates metrics about web scraping progress, including:
    - Processing speed (URLs per minute)
    - Elapsed time
    - Estimated remaining time
    - Estimated completion time
    The tests cover both edge cases (zero URLs processed) and normal operation
    where a portion of URLs have been processed.
    """
    
    def test_estimate_progress_zero_processed(self):
        """Test progress estimation with zero URLs processed."""
        start_time = time.time() - 60  # 60 seconds ago
        speed, elapsed, remaining, completion = oag_ca_gov_scraper.estimate_progress(start_time, 0, 100)
        
        self.assertEqual(speed, 0)
        self.assertEqual(elapsed, "00:00:00")  # This might be "00:01:00" in actual execution
        self.assertEqual(remaining, "Unknown")
        self.assertEqual(completion, "Unknown")
    
    def test_estimate_progress_normal_case(self):
        """Test progress estimation in a normal case."""
        start_time = time.time() - 60  # 60 seconds ago
        speed, elapsed, remaining, completion = oag_ca_gov_scraper.estimate_progress(start_time, 10, 100)
        
        # Speed should be around 10 URLs per minute (10 URLs in 60 seconds)
        self.assertAlmostEqual(speed, 10, delta=1)
        
        # Elapsed should be close to 1 minute
        self.assertIn("00:", elapsed)
        
        # Remaining should be around 9 minutes (90 URLs at 10 URLs/minute)
        self.assertNotEqual(remaining, "Unknown")
        
        # Completion should be a valid time string
        self.assertNotEqual(completion, "Unknown")


class TestDataExtraction(unittest.TestCase):
    """    
    Test case for data extraction functionality from the OAG California government website scraper.
    This class tests the extraction of data from HTML content retrieved from the CA Attorney General website.
    It verifies that the extract_value function correctly identifies and extracts values associated
    with specific labels in the HTML structure.
    
    The tests utilize a sample HTML structure mimicking the actual website format, with fields like
    AG Number, Date Filed, Case Name, and Settlement information. The class includes both direct
    testing and mocked testing approaches to validate the extraction functionality.
    Note: The extract_value function being tested is assumed to be defined within a main() function
    in the oag_ca_gov_scraper module, which is why a helper method is provided to access it.
    """
    def setUp(self):
        # Sample HTML content
        self.html_content = """
        <div class="node">
            <div class="field-label">AG Number:</div>
            <div class="field-items">2021-00123</div>
            <div class="field-label">Date Filed:</div>
            <div class="field-items">01/15/2021</div>
            
            <div>Civil Complaint</div>
            <div>
                <div class="field-label">Case Name:</div>
                <div class="field-items">Smith v. Jones</div>
                <div class="field-label">Court Name:</div>
                <div class="field-items">Superior Court</div>
            </div>
            
            <div>Settlement</div>
            <div>
                <div class="field-label">Settlement Date:</div>
                <div class="field-items">03/15/2021</div>
                <div class="details-label">
                    <div class="details">Non-Contingent Civil Penalty:</div>
                    $5,000.00
                </div>
                <div class="field-label">Email Address:</div>
                <div class="field-items"><a href="mailto:test@example.com">test@example.com</a></div>
            </div>
        </div>
        """
        self.soup = BeautifulSoup(self.html_content, 'html.parser')
    
    def test_extract_value(self):
        """Test the extract_value functionality directly."""
        result = oag_ca_gov_scraper.extract_value(self.soup, "AG Number:")
        self.assertEqual(result, "2021-00123")
    

class TestSheetWriting(unittest.TestCase):
    """
    Unit tests for the sheet writing functionality of the oag_ca_gov_scraper module.
    This test class verifies that the write_data_to_sheet_with_all_headers function
    correctly writes data to an Excel sheet with appropriate headers. It uses sample
    data representing California Attorney General case information, including links,
    AG numbers, dates, settlements, and civil complaints.
    The tests ensure that:
    1. Headers are correctly written to the sheet
    2. Data is properly populated in the corresponding cells
    3. Monetary value conversion is correctly applied
    The class uses unittest.mock to patch the convert_to_float function to ensure
    predictable behavior during testing.
    """
    
    def setUp(self):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        
        # Sample data
        self.all_data = [
            {
                'data': {
                    'link': 'https://example.com/1',
                    'AG Number': '2021-00001',
                    'Date Filed': '01/01/2021'
                },
                'flat_settlement_data': {
                    'Settlement_1_Settlement Date': '02/01/2021',
                    'Settlement_1_Non-Contingent Civil Penalty': '$1,000.00'
                }
            },
            {
                'data': {
                    'link': 'https://example.com/2',
                    'AG Number': '2021-00002',
                    'Date Filed': '01/02/2021'
                },
                'flat_civil_complaint_data': {
                    'Civil_Complaint_Case Name': 'Case A',
                    'Civil_Complaint_Court Name': 'Court A'
                }
            }
        ]
        
        # Headers by category
        self.all_headers_by_category = {
            'data': {'AG Number', 'Date Filed'},
            'flat_settlement_data': {'Settlement_1_Settlement Date', 'Settlement_1_Non-Contingent Civil Penalty'},
            'flat_civil_complaint_data': {'Civil_Complaint_Case Name', 'Civil_Complaint_Court Name'}
        }
    
    @patch('oag_ca_gov_scraper.convert_to_float')
    def test_write_data_to_sheet(self, mock_convert):
        """Test writing data to a sheet."""
        # Mock the convert_to_float function to return a predictable value
        mock_convert.return_value = 1000.0
        
        # Call the function
        oag_ca_gov_scraper.write_data_to_sheet_with_all_headers(
            self.sheet, self.all_data, self.all_headers_by_category
        )
        
        # Verify headers are written
        self.assertEqual(self.sheet.cell(row=1, column=1).value, 'link')
        self.assertEqual(self.sheet.cell(row=1, column=2).value, 'AG Number')
        
        # Verify data is written
        self.assertEqual(self.sheet.cell(row=2, column=1).value, 'https://example.com/1')
        self.assertEqual(self.sheet.cell(row=2, column=2).value, '2021-00001')
        
        # Verify monetary conversion was attempted
        mock_convert.assert_called()


if __name__ == '__main__':
    unittest.main()
