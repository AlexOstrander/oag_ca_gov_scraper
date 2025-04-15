# Import necessary libraries
import urllib2
from bs4 import BeautifulSoup
import xlwt
import pandas as pd  # Replace xlrd with pandas
import re
import sys
import os
# Add logging functionality
import logging
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import time  # Add this import at the top
import ssl
import math  # For isinf() function
import threading  # For multi-threading
import Queue  # For thread-safe queue in Python 2
import argparse  # For command-line argument parsing
import shutil  # For directory operations
from nile.utils.send_email import email_custom

# Thread-safe counter for tracking progress
class AtomicCounter:
    def __init__(self, initial=0):
        self.value = initial
        self.lock = threading.Lock()
        
    def increment(self):
        with self.lock:
            self.value += 1
            return self.value
            
    def get(self):
        with self.lock:
            return self.value
            
# Thread worker function to process URLs
def url_worker(url_queue, results_queue, failed_urls, counter, max_retries=3, retry_delay=5):
    """
    Worker function for thread pool to process URLs.
    
    Args:
        url_queue: Queue containing URLs to process
        results_queue: Queue to store processed results
        failed_urls: Shared list to track failed URLs
        counter: AtomicCounter to track progress
        max_retries: Maximum number of retry attempts
        retry_delay: Delay between retries in seconds
    """
    while not url_queue.empty():
        try:
            url = url_queue.get(block=False)
            thread_id = threading.current_thread().name
            
            # Track URL processing count for progress reporting
            current_count = counter.increment()
            
            retries = 0
            success = False
            
            while retries < max_retries and not success:
                try:
                    if retries > 0:
                        logging.info("Thread {}: Processing URL {} (Attempt {}/{})".format(
                            thread_id, url, retries+1, max_retries))
                    
                    # Create SSL context that ignores certificate validation
                    context = ssl._create_unverified_context()
                    response = urllib2.urlopen(url, context=context)
                    page_content = response.read()
                    
                    # Parse the page using BeautifulSoup
                    soup = BeautifulSoup(page_content, "html.parser")
                    
                    # Process the data (this will be defined in the main function)
                    data = process_url_data(url, soup)
                    
                    # Add to results queue
                    results_queue.put(data)
                    success = True
                    
                except urllib2.URLError as e:
                    retries += 1
                    error_msg = "Thread {}: Error fetching URL: {} - {} (Attempt {}/{})".format(
                        thread_id, url, str(e), retries, max_retries)
                    print(error_msg)
                    logging.error(error_msg)
                    
                    if retries < max_retries:
                        logging.info("Thread {}: Retrying in {} seconds...".format(thread_id, retry_delay))
                        time.sleep(retry_delay)
                    else:
                        with threading.Lock():
                            failed_urls.append(url)
                except Exception as e:
                    retries += 1
                    error_msg = "Thread {}: Error processing URL: {} - {} (Attempt {}/{})".format(
                        thread_id, url, str(e), retries, max_retries)
                    print(error_msg)
                    logging.error(error_msg)
                    
                    if retries < max_retries:
                        logging.info("Thread {}: Retrying in {} seconds...".format(thread_id, retry_delay))
                        time.sleep(retry_delay)
                    else:
                        with threading.Lock():
                            failed_urls.append(url)
                
            url_queue.task_done()
            
        except Queue.Empty:
            break  # Queue is empty, exit the worker
        except Exception as e:
            logging.error("Thread {}: Unexpected error: {}".format(thread_id, str(e)))
            break

# Function to process URL data (will be called by worker threads)
def process_url_data(url, soup):
    """
    Process the data from a URL.
    
    Args:
        url: The URL being processed
        soup: BeautifulSoup object containing the parsed HTML
        
    Returns:
        Dictionary containing the extracted data
    """
    # Check if notice has been withdrawn
    notice_withdrawn = False
    withdrawal_data = {}

    # Look for the withdrawal banner
    withdrawal_banner = soup.find("span", class_="label-danger", 
                                string=lambda x: x and "THIS 60-DAY NOTICE HAS BEEN WITHDRAWN" in x)
    if withdrawal_banner:
        notice_withdrawn = True
    
    # Extract Withdrawal Letter
    withdrawal_letter_div = soup.find("div", class_="field-label", string=lambda x: x and "Withdrawal Letter:" in x)
    if withdrawal_letter_div and withdrawal_letter_div.find_next_sibling():
        letter_item = withdrawal_letter_div.find_next_sibling().find("a")
        if letter_item and letter_item.get("href"):
            pdf_url = letter_item.get("href")
            pdf_name = letter_item.text.strip()
            withdrawal_data["Withdrawal Letter"] = "[{}]({})".format(pdf_name, pdf_url)
            withdrawal_data["Withdrawal Status"] = "Withdrawn" 
            withdrawal_data["Withdrawal ID"] = url.split("/")[-1]
            withdrawal_data["Withdrawal Date"] = extract_value(soup, "Withdrawal Date:")

    # Extract AG Number and ensure it's properly formatted
    ag_number = extract_value(soup, "AG Number:")
    
    # Extract the year and number parts from the URL
    url_parts = url.split("-")
    if len(url_parts) >= 2:
        year_part = url_parts[-2]
        number_part = url_parts[-1]
        
        # Format the number part to ensure 5 digits
        formatted_number = format_ag_number(number_part)
        
        # Format the PDF URL with the properly formatted number
        pdf_url = "https://oag.ca.gov/prop65/60-Day-Notice-{}-{}/{}.pdf".format(
            year_part, formatted_number, formatted_number)
    else:
        # Fallback if URL structure doesn't match expected format
        url_id = url.split("/")[-1]
        pdf_url = "https://oag.ca.gov/prop65/60-Day-Notice-{0}/{0}.pdf".format(url_id)
    
    # Extract relevant data from the main section
    main_data = {
        "link": url,
        "AG Number": ag_number,
        "Notice PDF": "[{}.pdf]({})".format(url.split("/")[-1], pdf_url),
        "Date Filed": extract_value(soup, "Date Filed:"),
        "Noticing Party": extract_value(soup, "Noticing Party:"),
        "Plaintiff Attorney": extract_value(soup, "Plaintiff Attorney:"),
        "Alleged Violators": extract_value(soup, "Alleged Violators:"),
        "Chemicals": extract_value(soup, "Chemicals:"),
        "Source": extract_value(soup, "Source:"),
    }

    # Add withdrawal information if notice has been withdrawn
    if notice_withdrawn:
        main_data.update(withdrawal_data)

    # Extract Civil Complaint Data
    civil_complaint_div = soup.find("div", text="Civil Complaint")
    flat_civil_complaint_data = {}
    if civil_complaint_div:
        civil_complaint_data = extract_section_data(civil_complaint_div, "Civil Complaint")
        for key, value in civil_complaint_data.items():
            flat_civil_complaint_data["Civil_Complaint_{}".format(key)] = value

    # Find all settlement-related divs
    all_settlement_divs = soup.find_all("div", string=lambda s: s and s.strip() in ["Settlement", "Corrected Settlement"])
    
    # Separate into correct categories
    corrected_settlement_divs = [div for div in all_settlement_divs if div.string and div.string.strip() == "Corrected Settlement"]
    settlement_divs = [div for div in all_settlement_divs if div.string and div.string.strip() == "Settlement"]
    
    # Extract Corrected Settlement Data (up to 5 settlements)
    flat_corrected_settlement_data = {}
    for i, div in enumerate(corrected_settlement_divs[:5]):
        data = extract_section_data(div, "Settlement")
        for key, value in data.items():
            flat_corrected_settlement_data["Corrected_Settlement_{}_{}".format(i+1, key)] = value or ""

    # Extract Settlement Data (up to 5 settlements)
    flat_settlement_data = {}
    for i, div in enumerate(settlement_divs[:5]):
        data = extract_section_data(div, "Settlement")
        for key, value in data.items():
            flat_settlement_data["Settlement_{}_{}".format(i+1, key)] = value or ""

    # Extract Judgment Data (up to 5 judgments)
    judgment_divs = soup.find_all("div", text="Judgment")
    flat_judgment_data = {}
    for i, div in enumerate(judgment_divs[:5]):
        data = extract_section_data(div, "Judgment")
        for key, value in data.items():
            flat_judgment_data["Judgment_{}_{}".format(i+1, key)] = value or ""

    # Organize data into a dictionary with specific order
    organized_data = {'data': main_data}
    
    if flat_civil_complaint_data:
        organized_data['flat_civil_complaint_data'] = flat_civil_complaint_data
        
    if flat_settlement_data:
        organized_data['flat_settlement_data'] = flat_settlement_data
        
    if flat_judgment_data:
        organized_data['flat_judgment_data'] = flat_judgment_data

    if flat_corrected_settlement_data:
        organized_data['flat_corrected_settlement_data'] = flat_corrected_settlement_data

    return organized_data

def read_urls_from_excel(excel_path, column_index=0, sheet_index=0):
    """
    Read URLs from an Excel file using pandas.
    Args:
        excel_path: Path to the Excel file
        column_index: Index of the column containing URLs (default: 0, i.e., first column)
        sheet_index: Index of the sheet containing URLs (default: 0, i.e., first sheet)
    Returns:
        List of URLs
    """
    try:
        # Read Excel file using pandas
        df = pd.read_excel(excel_path, sheet_name=sheet_index)
        
        # Get the column name (or use the index if it's a number)
        if isinstance(column_index, int):
            column_name = df.columns[column_index]
        else:
            column_name = column_index
            
        # Extract URLs from the specified column
        urls = []
        for value in df[column_name].dropna():
            # Convert to string in case it's not already
            value_str = str(value).strip()
            if value_str.startswith("http"):
                urls.append(value_str)
        
        return urls
    except Exception as e:
        print("Error reading Excel file: {}".format(e))
        return []

# Add a new function to read from TSV files
def read_urls_from_tsv(tsv_path, column_index=0):
    """
    Read URLs from a TSV file.
    Args:
        tsv_path: Path to the TSV file
        column_index: Index of the column containing URLs (default: 0, i.e., first column)
    Returns:
        List of URLs
    """
    try:
        urls = []
        with open(tsv_path, 'r') as f:
            # Skip header line if present
            next(f, None)
            for line in f:
                parts = line.strip().split('\t')
                if len(parts) > column_index:
                    url = parts[column_index].strip()
                    if url.startswith("http"):
                        urls.append(url)
        return urls
    except Exception as e:
        print("Error reading TSV file: {}".format(e))
        return []


# Add this helper function to convert monetary values to floats
def convert_to_float(value_str):
    """Convert a monetary string (like $1,234.56) to a float (1234.56)"""
    if value_str is None or value_str == '':
        return ""
    
    try:
        # Handle values that are already numbers
        if isinstance(value_str, (int, float)):
            return float(value_str)
            
        # Remove dollar signs, commas, and spaces
        cleaned_str = str(value_str).replace('$', '').replace(',', '').replace(' ', '')
        # If there's nothing left, return empty string
        if not cleaned_str:
            return ""
        # Convert to float    
        return float(cleaned_str)
    except (ValueError, TypeError):
        # If conversion fails, return original string
        return value_str


def generate_urls_for_year_range(start_year, end_year, start_ids, end_ids):
    """
    Generate a list of all possible URLs for the given year range.
    
    Args:
        start_year: Starting year (e.g., 2015)
        end_year: Ending year (e.g., 2025)
        start_ids: Dictionary mapping years to starting IDs
        end_ids: Dictionary mapping years to ending IDs
        
    Returns:
        List of URLs
    """
    urls = []
    
    for year in range(start_year, end_year + 1):
        if year in start_ids:
            # Extract the numeric part of the ID and ensure they're integers
            start_id = int(start_ids[year])
            
            # Get the end_id - we'll use this as a starting point but may exceed it
            end_id = None
            if year in end_ids:
                end_id = int(end_ids[year])
            else:
                # If no end_id is specified, use a large number for the current year
                # and a reasonable limit for past years
                current_year = datetime.datetime.now().year
                if year == current_year:
                    end_id = 10000  # A large number for the current year
                else:
                    end_id = 5000   # A reasonable limit for past years
            
            print("Processing year {} from ID {} to at least {}".format(year, start_id, end_id))
            
            # Format with leading zeros to ensure 5 digits (fixed format)
            id_format = "{:05d}"  # Always 5 digit format
            
            # Generate URLs for the specified range
            for id_num in range(start_id, end_id + 1):
                formatted_id = id_format.format(id_num)
                url = "https://oag.ca.gov/prop65/60-Day-Notice-{}-{}".format(year, formatted_id)
                urls.append(url)
                
    return urls

# Function to dynamically discover the end_id for a year by testing URLs
def discover_year_end_id(year, start_id):
    """
    Dynamically discover the last valid ID for a given year by testing URLs
    until 3 consecutive failures are encountered.
    
    Args:
        year: The year to check
        start_id: The ID to start checking from
        
    Returns:
        The highest valid ID found
    """
    id_format = "{:05d}"  # Always 5 digit format
    current_id = start_id
    consecutive_failures = 0
    max_consecutive_failures = 3
    last_valid_id = start_id - 1  # Initialize to one less than start
    
    print("Discovering end ID for year {}, starting from {}...".format(year, start_id))
    
    while consecutive_failures < max_consecutive_failures:
        formatted_id = id_format.format(current_id)
        url = "https://oag.ca.gov/prop65/60-Day-Notice-{}-{}".format(year, formatted_id)
        
        try:
            # Try to access the URL
            print("Testing URL: {}".format(url))
            context = ssl._create_unverified_context()
            response = urllib2.urlopen(url, context=context)
            
            # If successful, update the last valid ID and reset failure counter
            last_valid_id = current_id
            consecutive_failures = 0
            print("Found valid ID: {}".format(current_id))
            
        except urllib2.HTTPError as e:
            if e.code == 404:
                # URL not found - increment failure counter
                consecutive_failures += 1
                print("ID {} not found ({} consecutive failures)".format(
                    current_id, consecutive_failures))
            else:
                # Other HTTP error - treat as temporary failure
                print("HTTP error {} for ID {}, treating as temporary".format(e.code, current_id))
                time.sleep(1)  # Pause to avoid overwhelming the server
                
        except Exception as e:
            # Other error - treat as temporary failure
            print("Error checking ID {}: {}".format(current_id, str(e)))
            time.sleep(1)  # Pause to avoid overwhelming the server
            
        # Increment the ID
        current_id += 1
        
        # Small pause to be nice to the server
        time.sleep(0.5)
    
    print("Discovered end ID for year {}: {}".format(year, last_valid_id))
    return last_valid_id

def auto_discover_urls_for_year_range(start_year, end_year, start_ids):
    """
    Generate a list of all possible URLs for the given year range,
    automatically discovering the end IDs for each year.
    
    Args:
        start_year: Starting year (e.g., 2015)
        end_year: Ending year (e.g., 2025)
        start_ids: Dictionary mapping years to starting IDs
        
    Returns:
        List of URLs
    """
    urls = []
    discovered_end_ids = {}
    
    for year in range(start_year, end_year + 1):
        if year in start_ids:
            # Extract the numeric part of the ID and ensure they're integers
            start_id = int(start_ids[year])
            
            # Get the current known end ID as a starting point for discovery
            current_known_end = None
            if year in end_ids:
                current_known_end = int(end_ids[year])
            else:
                # If no end_id is specified, use a reasonable starting point
                current_known_end = start_id + 500
            
            # Discover the actual end ID
            end_id = discover_year_end_id(year, current_known_end)
            discovered_end_ids[year] = end_id
            
            # Format with leading zeros to ensure 5 digits (fixed format)
            id_format = "{:05d}"  # Always 5 digit format
            
            # Generate URLs for the specified range
            for id_num in range(start_id, end_id + 1):
                formatted_id = id_format.format(id_num)
                url = "https://oag.ca.gov/prop65/60-Day-Notice-{}-{}".format(year, formatted_id)
                urls.append(url)
    
    # Print a summary of what we discovered
    print("\nDiscovered End IDs:")
    for year in sorted(discovered_end_ids.keys()):
        print("  {}: {}".format(year, discovered_end_ids[year]))
                
    return urls

# Example usage with proper 5-digit formatting:
start_ids = {
    2015: 1,      # Will be formatted as 00001
    2016: 1,      # Will be formatted as 00001
    2017: 1,      # Will be formatted as 00001
    2018: 1,      # Will be formatted as 00001
    2019: 1,      # Will be formatted as 00001
    2020: 1,      # Will be formatted as 00001
    2021: 1,      # Will be formatted as 00001
    2022: 1,      # Will be formatted as 00001
    2023: 1,      # Will be formatted as 00001
    2024: 1,      # Will be formatted as 00001
    2025: 1       # Will be formatted as 00001
}

end_ids = {
    2015: 1349,   # Will be formatted as 01349
    2016: 1582,   # Will be formatted as 01582
    2017: 2714,   # Will be formatted as 02714
    2018: 2369,   # Will be formatted as 02369
    2019: 2425,   # Will be formatted as 02425
    2020: 3574,   # Will be formatted as 03574
    2021: 3166,   # Will be formatted as 03166
    2022: 3175,   # Will be formatted as 03175
    2023: 4144,   # Will be formatted as 04144
    2024: 5404,   # Will be formatted as 05404
    2025: 1211    # Will be formatted as 02000
}

# Generate URLs for the past 10 years
all_urls = generate_urls_for_year_range(2015, 2025, start_ids, end_ids)


# Function to format AG numbers properly with 5 digits
def format_ag_number(ag_number):
    """Format AG number with leading zeros to ensure 5 digits"""
    if not ag_number:
        return "00000"
    try:
        # Extract numeric portion if it contains hyphens or other characters
        numeric_part = "".join([c for c in str(ag_number) if c.isdigit()])
        if numeric_part:
            return "{:05d}".format(int(numeric_part))
    except (ValueError, TypeError):
        pass
    return str(ag_number).zfill(5)  # fallback
    
# Move this function outside main()
def write_data_to_sheet_with_all_headers(sheet, all_data, all_headers_by_category):
    # Define data categories here to ensure they're in scope
    data_categories = ['data', 'flat_civil_complaint_data', 'flat_settlement_data', 
                    'flat_judgment_data', 'flat_corrected_settlement_data']
    
    # Define the exact column order with withdrawal columns removed
    ordered_headers = [
        # Main data
        'link',
        'AG Number',
        'Alleged Violators',
        'Chemicals',
        'Date Filed',
        'Notice PDF',
        'Noticing Party',
        'Plaintiff Attorney',
        'Source',
        
        # Only include Withdrawal Status and ID
        'Withdrawal Status',
        'Withdrawal ID',
        'Withdrawal Date',
        'Withdrawal Letter',
        
        # Civil Complaint data
        'Civil_Complaint_Date Filed',
        'Civil_Complaint_Case Name',
        'Civil_Complaint_Court Name',
        'Civil_Complaint_Court Docket Number',
        'Civil_Complaint_Plaintiff',
        'Civil_Complaint_Plaintiff Attorney',
        'Civil_Complaint_Defendant',
        'Civil_Complaint_Type of Claim',
        'Civil_Complaint_Relief Sought',
        'Civil_Complaint_Contact Name',
        'Civil_Complaint_Contact Organization',
        'Civil_Complaint_Email Address',
        'Civil_Complaint_Address',
        'Civil_Complaint_City, State, Zip',
        'Civil_Complaint_Phone Number'
    ]

    # Add Settlement columns (1-5) instead of just (1-3)
    for settlement_num in range(1, 6):
        settlement_fields = [
            'Settlement_{}_Settlement Date'.format(settlement_num),
            'Settlement_{}_Case Name'.format(settlement_num),
            'Settlement_{}_Court Name'.format(settlement_num),
            'Settlement_{}_Court Docket Number'.format(settlement_num),
            'Settlement_{}_Plaintiff'.format(settlement_num),
            'Settlement_{}_Plaintiff Attorney'.format(settlement_num),
            'Settlement_{}_Defendant'.format(settlement_num),
            'Settlement_{}_Injunctive Relief'.format(settlement_num),
            'Settlement_{}_Non-Contingent Civil Penalty'.format(settlement_num),
            'Settlement_{}_Attorneys Fees and Costs'.format(settlement_num),
            'Settlement_{}_Payment in Lieu of Penalty'.format(settlement_num),
            'Settlement_{}_Total Payments'.format(settlement_num),
            'Settlement_{}_Will settlement be submitted to court?'.format(settlement_num),
            'Settlement_{}_Contact Name'.format(settlement_num),
            'Settlement_{}_Contact Organization'.format(settlement_num),
            'Settlement_{}_Email Address'.format(settlement_num),
            'Settlement_{}_Address'.format(settlement_num),
            'Settlement_{}_City, State, Zip'.format(settlement_num),
            'Settlement_{}_Phone Number'.format(settlement_num)
        ]
        ordered_headers.extend(settlement_fields)
    
    # Add Corrected Settlement columns (1-5) instead of just (1-3)
    for settlement_num in range(1, 6):
        corrected_settlement_fields = [
            'Corrected_Settlement_{}_Settlement Date'.format(settlement_num),
            'Corrected_Settlement_{}_Case Name'.format(settlement_num),
            'Corrected_Settlement_{}_Court Name'.format(settlement_num),
            'Corrected_Settlement_{}_Court Docket Number'.format(settlement_num),
            'Corrected_Settlement_{}_Plaintiff'.format(settlement_num),
            'Corrected_Settlement_{}_Plaintiff Attorney'.format(settlement_num),
            'Corrected_Settlement_{}_Defendant'.format(settlement_num),
            'Corrected_Settlement_{}_Injunctive Relief'.format(settlement_num),
            'Corrected_Settlement_{}_Non-Contingent Civil Penalty'.format(settlement_num),
            'Corrected_Settlement_{}_Attorneys Fees and Costs'.format(settlement_num),
            'Corrected_Settlement_{}_Payment in Lieu of Penalty'.format(settlement_num),
            'Corrected_Settlement_{}_Total Payments'.format(settlement_num),
            'Corrected_Settlement_{}_Will settlement be submitted to court?'.format(settlement_num),
            'Corrected_Settlement_{}_Contact Name'.format(settlement_num),
            'Corrected_Settlement_{}_Contact Organization'.format(settlement_num),
            'Corrected_Settlement_{}_Email Address'.format(settlement_num),
            'Corrected_Settlement_{}_Address'.format(settlement_num),
            'Corrected_Settlement_{}_City, State, Zip'.format(settlement_num),
            'Corrected_Settlement_{}_Phone Number'.format(settlement_num)
        ]
        ordered_headers.extend(corrected_settlement_fields)
    
    # Add Judgment columns (1-5) instead of just 1
    for judgment_num in range(1, 6):
        judgment_fields = [
            'Judgment_{}_Judgment Date'.format(judgment_num),
            'Judgment_{}_Settlement reported to AG'.format(judgment_num),
            'Judgment_{}_Case Name'.format(judgment_num),
            'Judgment_{}_Court Name'.format(judgment_num),
            'Judgment_{}_Court Docket Number'.format(judgment_num),
            'Judgment_{}_Plaintiff'.format(judgment_num),
            'Judgment_{}_Plaintiff Attorney'.format(judgment_num),
            'Judgment_{}_Defendant'.format(judgment_num),
            'Judgment_{}_Injunctive Relief'.format(judgment_num),
            'Judgment_{}_Non-Contingent Civil Penalty'.format(judgment_num),
            'Judgment_{}_Attorneys Fees and Costs'.format(judgment_num),
            'Judgment_{}_Payment in Lieu of Penalty'.format(judgment_num),
            'Judgment_{}_Total Payments'.format(judgment_num),
            'Judgment_{}_Is Judgment Pursuant to Settlement?'.format(judgment_num),
            'Judgment_{}_Contact Name'.format(judgment_num),
            'Judgment_{}_Contact Organization'.format(judgment_num),
            'Judgment_{}_Email Address'.format(judgment_num),
            'Judgment_{}_Address'.format(judgment_num),
            'Judgment_{}_City, State, Zip'.format(judgment_num),
            'Judgment_{}_Phone Number'.format(judgment_num)
        ]
        ordered_headers.extend(judgment_fields)
    
    # Write headers in the first row - openpyxl uses 1-based indexing
    for col_idx, header in enumerate(ordered_headers, 1):
        sheet.cell(row=1, column=col_idx, value=header)
    
    # Create a mapping of headers to their column index - openpyxl uses 1-based indexing
    header_to_column = {header: i for i, header in enumerate(ordered_headers, 1)}
    
    # Initialize all empty values to prevent None errors
    for col_idx, _ in enumerate(ordered_headers, 1):
        for row_idx in range(2, len(all_data) + 2):
            sheet.cell(row=row_idx, column=col_idx, value="")
    
    # Process each URL's data - openpyxl uses 1-based indexing for rows too
    row_idx = 2  # Start from row 2 (after headers)
    
    def is_monetary_field(header):
        monetary_keywords = [
            "civil penalty", 
            "fees", 
            "costs",
            "payment",
            "penalty",
            "payments",
            "total"
        ]
        header_lower = header.lower()
        return any(keyword in header_lower for keyword in monetary_keywords)

    for entry in all_data:
        # Process all categories of data
        for category in data_categories:
            if category in entry:
                for header, value in entry[category].items():
                    # Find the column for this header
                    if header in header_to_column:
                        # Check if this is a monetary field that should be converted to float
                        if is_monetary_field(header):
                            value = convert_to_float(value)
                        
                        # Write value to cell - only if not empty
                        if value:  # Only write non-empty values
                            sheet.cell(row=row_idx, column=header_to_column[header], value=value)
        
        # Handle 'link' value separately to avoid duplication
        if 'data' in entry and 'link' in entry.get('data', {}):
            sheet.cell(row=row_idx, column=header_to_column['link'], 
                    value=entry['data']['link'])
        
        # Move to next row
        row_idx += 1

# Function to safely extract data
def extract_value(soup, label_text):
    label = soup.find("div", class_="field-label", string=lambda x: x and label_text in x)
    if label and label.find_next_sibling():
        sibling = label.find_next_sibling()
        if sibling:
            return sibling.text.strip()
    return ""

def extract_value_from_element(element, label_text):
    children = element.find_next()
    if not children:
        return ""

    # Special handling for Non-Contingent Civil Penalty
    if label_text == 'Non-Contingent Civil Penalty:':
        # Look for details-label div with details div containing Non-Contingent Civil Penalty text
        penalty_fields = children.find_all('div', class_='details-label')
        for field in penalty_fields:
            details = field.find('div', class_='details')
            if details and 'Non-Contingent Civil Penalty' in details.text:
                # Extract the text directly from the details-label div (outside the details div)
                # First get all text in the parent div
                full_text = field.get_text().strip()
                # Remove the text from the nested details div
                details_text = details.get_text().strip()
                # What remains should be the amount
                amount = full_text.replace(details_text, '').strip()
                return amount

    # Special handling for Address fields
    if label_text == 'Address:':
        # Try to find field with prop65-address class first
        address_field = children.find('div', class_=lambda c: c and "field-name-field-prop65-address" in c)
        if address_field:
            field_item = address_field.find('div', class_='field-item')
            if field_item:
                return field_item.text.strip()
                
        # Try to find div containing the address label
        address_label = children.find(lambda tag: tag.name == 'div' and 
                                    (tag.get('class') and 'field-label' in tag.get('class')) and 
                                    label_text in tag.text)
        if address_label and address_label.find_next_sibling():
            return address_label.find_next_sibling().text.strip()
    
    # Special handling for Email Address
    if label_text == 'Email Address:':
        # Try to find a link (a tag) which would indicate an email address
        email_a_tag = children.find('a', href=lambda href: href and 'mailto:' in href)
        if email_a_tag:
            return email_a_tag.text.strip()
        
        # If no mailto link is found, try the standard field-label approach
        email_div = children.find('div', class_="field-label", string=lambda x: x and label_text in x)
        if email_div and email_div.find_next_sibling():
            return email_div.find_next_sibling().text.strip()
        
        # If still not found, try the details class
        email_div = children.find('div', class_="details", string=lambda x: x and label_text in x)
        if email_div:
            # Look for an anchor tag that might contain the email
            email_a = email_div.parent.find('a')
            if email_a:
                return email_a.text.strip()
            
            # If no anchor tag, try getting text after the details div
            if email_div.parent:
                next_text = email_div.parent.contents[-1].strip() if len(email_div.parent.contents) > 1 else ""
                if next_text:
                    return next_text
                
        return ""
    
    # Try to find field-label first (general case)
    sibling = children.find('div', class_="field-label", string=lambda x: x and label_text in x)
    if sibling and sibling.next_sibling:
        return sibling.next_sibling.text.strip()
    
    # If not found, try details class
    sibling = children.find('div', class_="details", string=lambda x: x and label_text in x)
    if sibling:
        # Get the parent of the details div
        parent = sibling.parent
        if parent:
            # Get the text directly after the div.details
            next_text = parent.contents[-1].strip() if len(parent.contents) > 1 else ""
            if next_text:
                return next_text
                
        # Try other potential locations
        next_text = sibling.next_sibling.strip() if sibling.next_sibling else ""
        if next_text:
            return next_text
            
        field_item = sibling.find_next('div', class_='field-item')
        if field_item:
            return field_item.text.strip()
            
    return ""

# Consolidated function to extract data based on section type
def extract_section_data(div, section_type):
    data = {}
    field_mapping = {
        "Civil Complaint": {
            "Case Name": "Case Name:",
            "Court Name": "Court Name:",
            "Date Filed": "Date Filed:",
            "Court Docket Number": "Court Docket Number:",
            "Plaintiff": "Plaintiff:",
            "Plaintiff Attorney": "Plaintiff Attorney:",
            "Defendant": "Defendant:",
            "Type of Claim": "Type of Claim:",
            "Relief Sought": "Relief Sought:",
            "Contact Name": "Contact Name:",
            "Contact Organization": "Contact Organization:",
            "Email Address": "Email Address:",
            "Address": "Address:",
            "City, State, Zip": "City, State, Zip:",
            "Phone Number": "Phone Number:"
        },
        "Settlement": {
            "Settlement Date": "Settlement Date:",
            "Case Name": "Case Name:",
            "Court Name": "Court Name:",
            "Court Docket Number": "Court Docket Number:",
            "Plaintiff": "Plaintiff:",
            "Plaintiff Attorney": "Plaintiff Attorney:",
            "Defendant": "Defendant:",
            "Injunctive Relief": "Injunctive Relief:",
            "Non-Contingent Civil Penalty": "Non-Contingent Civil Penalty:",
            "Attorneys Fees and Costs": "Attorney(s) Fees and Costs:",
            "Payment in Lieu of Penalty": "Payment in Lieu of Penalty:",
            "Total Payments": "Total Payments:",
            "Will settlement be submitted to court?": "Will settlement be submitted to court?",
            "Contact Name": "Contact Name:",
            "Contact Organization": "Contact Organization:",
            "Email Address": "Email Address:",
            "Address": "Address:",
            "City, State, Zip": "City, State, Zip:",
            "Phone Number": "Phone Number:"
        },
        "Judgment": {
            "Judgment Date": "Judgment Date:",
            "Settlement reported to AG": "Settlement reported to AG:",
            "Case Name": "Case Name:",
            "Court Name": "Court Name:",
            "Court Docket Number": "Court Docket Number:",
            "Plaintiff": "Plaintiff:",
            "Plaintiff Attorney": "Plaintiff Attorney:",
            "Defendant": "Defendant:",
            "Injunctive Relief": "Injunctive Relief:",
            "Non-Contingent Civil Penalty": "Non-Contingent Civil Penalty:",
            "Attorneys Fees and Costs": "Attorney(s) Fees and Costs:",
            "Payment in Lieu of Penalty": "Payment in Lieu of Penalty:",
            "Total Payments": "Total Payments:",
            "Is Judgment Pursuant to Settlement?": "Is Judgment Pursuant to Settlement?",
            "Contact Name": "Contact Name:",
            "Contact Organization": "Contact Organization:",
            "Email Address": "Email Address:",
            "Address": "Address:",
            "City, State, Zip": "City, State, Zip:",
            "Phone Number": "Phone Number:"
        }
    }
    
    mapping = field_mapping.get(section_type, field_mapping["Settlement"])
    
    for field, label in mapping.items():
        value = extract_value_from_element(div, label)
        data[field] = value if value is not None else ""
    
    return data

# Update the main function to handle TSV files
def main(year_range=None, compare_file=None, input_file=None, threads=5, no_auto_discover=False):
    """
    Main function to execute the OAG CA Gov scraper.
    This function sets up logging, processes input files or URLs, and scrapes data from the provided URLs.
    It supports reading URLs from Excel (.xls, .xlsx) or TSV (.tsv, .txt) files, or using a hardcoded list of URLs.
    The scraped data is then saved to an Excel file.
    
    Args:
        year_range (tuple): Optional tuple of (start_year, end_year) for scraping
        compare_file (str): Optional path to Excel file with existing data to compare against
        input_file (str): Optional path to input file with URLs to scrape (.xls, .xlsx, .tsv, or .txt)
        threads (int): Number of threads to use for scraping (default: 5)
        no_auto_discover (bool): If True, disables auto-discovery of new records
    """
    # For script usage, parse command line arguments if no parameters provided
    if all(param is None for param in (year_range, compare_file, input_file)) and len(sys.argv) > 1:
        parser = argparse.ArgumentParser(description='OAG CA Gov Prop65 Notice Scraper')
        parser.add_argument('--year-range', nargs=2, metavar=('START_YEAR', 'END_YEAR'),
                            type=int, help='Specify start and end year for scraping')
        parser.add_argument('--compare-file', metavar='FILE', type=str,
                            help='Excel file with existing data to compare against')
        parser.add_argument('--input-file', metavar='FILE', type=str,
                            help='Input file with URLs to scrape (.xls, .xlsx, .tsv, or .txt)')
        parser.add_argument('--threads', type=int, default=5,
                            help='Number of threads to use for scraping (default: 5)')
        parser.add_argument('--no-auto-discover', action='store_true',
                            help='Disable auto-discovery of new records')
        args = parser.parse_args()
        
        # Use command line args if provided
        year_range = args.year_range
        compare_file = args.compare_file
        input_file = args.input_file
        threads = args.threads
        no_auto_discover = args.no_auto_discover
    
    # Create reports directory if it doesn't exist
    reports_dir = "/reports/oag_gov_scrapes"
    if not os.path.exists(reports_dir):
        try:
            os.makedirs(reports_dir)
        except OSError as e:
            print("Error creating reports directory: {}".format(e))
            reports_dir = "."  # Use current directory as fallback
    
    # Set up logging to file with date-only timestamp (YYYY_MM_DD)
    date_stamp = datetime.datetime.now().strftime("%Y_%m_%d")
    log_filename = os.path.join(reports_dir, "scraper_errors_{}.log".format(date_stamp))
    logging.basicConfig(
        filename=log_filename,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    
    # Log start of scraping
    logging.info("Starting OAG CA Gov scraper")
    
    # Initialize variables that might be used later
    url_queue = Queue.Queue()
    results_queue = Queue.Queue()
    failed_urls = []
    counter = AtomicCounter()
    num_threads = threads  # Use provided threads value
    
    # Check for comparison file first - this needs to happen before determining URLs
    comparison_data = None
    urls_from_comparison = []
    if compare_file and os.path.exists(compare_file):
        print("Loading comparison data from: {}".format(compare_file))
        logging.info("Loading comparison data from: {}".format(compare_file))
        try:
            comparison_data = load_comparison_data(compare_file)
            if comparison_data:
                print("Found {} existing records to check".format(len(comparison_data)))
                logging.info("Found {} existing records to check".format(len(comparison_data)))
                
                # Extract URLs from the comparison file
                urls_from_comparison = list(comparison_data.keys())
                print("Extracted {} URLs from comparison file to check for updates".format(len(urls_from_comparison)))
                logging.info("Extracted {} URLs from comparison file to check for updates".format(len(urls_from_comparison)))
                
                # When using auto-discovery, also check for new records at the end of each year
                if not no_auto_discover:
                    print("Auto-discovery enabled - will also check for new records")
                    # Extract years from URLs to auto-discover
                    years_to_check = set()
                    for url in urls_from_comparison:
                        url_parts = url.split("-")
                        if len(url_parts) >= 2:
                            try:
                                year = int(url_parts[-2])
                                years_to_check.add(year)
                            except (ValueError, IndexError):
                                pass
                    
                    if years_to_check:
                        print("Years found in comparison file: {}".format(sorted(years_to_check)))
                        # For each year, find the highest ID in the existing URLs
                        highest_ids = {}
                        for year in years_to_check:
                            highest_id = 0
                            for url in urls_from_comparison:
                                if "-{}-".format(year) in url:
                                    url_parts = url.split("-")
                                    if len(url_parts) >= 2:
                                        try:
                                            id_num = int(url_parts[-1])
                                            highest_id = max(highest_id, id_num)
                                        except (ValueError, IndexError):
                                            pass
                            if highest_id > 0:
                                highest_ids[year] = highest_id
                        
                        # Auto-discover new URLs for each year starting from the highest known ID
                        auto_discovered_urls = []
                        for year in sorted(highest_ids.keys()):
                            print("Auto-discovering new URLs for year {} starting from ID {}".format(
                                year, highest_ids[year] + 1))
                            # Start looking from one past the highest known ID
                            end_id = discover_year_end_id(year, highest_ids[year] + 1)
                            if end_id > highest_ids[year]:
                                print("Found {} new records for year {}".format(
                                    end_id - highest_ids[year], year))
                                # Generate URLs for the new range
                                for id_num in range(highest_ids[year] + 1, end_id + 1):
                                    formatted_id = "{:05d}".format(id_num)
                                    url = "https://oag.ca.gov/prop65/60-Day-Notice-{}-{}".format(year, formatted_id)
                                    auto_discovered_urls.append(url)
                            else:
                                print("No new records found for year {}".format(year))
                        
                        if auto_discovered_urls:
                            print("Auto-discovered {} new URLs to check".format(len(auto_discovered_urls)))
                            urls_from_comparison.extend(auto_discovered_urls)
                
                # If we have a comparison file, we'll ONLY process those URLs
                if urls_from_comparison:
                    urls = urls_from_comparison
                    print("Will process {} URLs from comparison file".format(len(urls)))
                    logging.info("Will process {} URLs from comparison file".format(len(urls)))
                    
                    # Create queues for multi-threading
                    url_queue = Queue.Queue()
                    results_queue = Queue.Queue()
                    
                    # Add URLs to the queue
                    for url in urls:
                        url_queue.put(url)
                    
                    # Shared list to track failed URLs (thread-safe with locks)
                    failed_urls = []
                    
                    # Create a counter for tracking progress
                    counter = AtomicCounter()
                    
                    # Determine the number of threads to use
                    num_threads = min(threads, len(urls))
                    print("Using {} threads for scraping".format(num_threads))
                    logging.info("Using {} threads for scraping".format(num_threads))
                    
                    # Skip the rest of the URL determination logic
                    goto_start_scraping = True
                else:
                    goto_start_scraping = False
                    print("No valid URLs found in comparison file, will use regular URL sources")
                    logging.info("No valid URLs found in comparison file, will use regular URL sources")
            else:
                goto_start_scraping = False
                print("No valid comparison data found, will use regular URL sources")
                logging.info("No valid comparison data found, will use regular URL sources")
        except Exception as e:
            goto_start_scraping = False
            print("Error loading comparison data: {}".format(e))
            logging.error("Error loading comparison data: {}".format(e))
    else:
        goto_start_scraping = False
    
    # Only determine URLs if we're not using comparison file URLs
    if not goto_start_scraping:
        # Determine which URLs to scrape based on command line arguments
        urls = []
        
        # Option 1: Check for input file (highest priority)
        if input_file and os.path.exists(input_file):
            input_file_path = input_file
            
            # Check if it's an Excel file
            if input_file_path.endswith(('.xls', '.xlsx')):
                print("Reading URLs from Excel file: {}".format(input_file_path))
                logging.info("Reading URLs from Excel file: {}".format(input_file_path))
                urls = read_urls_from_excel(input_file_path)
                
            # Check if it's a TSV file
            elif input_file_path.endswith('.tsv') or input_file_path.endswith('.txt'):
                print("Reading URLs from TSV file: {}".format(input_file_path))
                logging.info("Reading URLs from TSV file: {}".format(input_file_path))
                urls = read_urls_from_tsv(input_file_path)
                
            else:
                error_msg = "Unsupported file format. Please provide an Excel (.xls, .xlsx) or TSV (.tsv, .txt) file."
                print(error_msg)
                logging.error(error_msg)
                return
                
            if not urls:
                error_msg = "No valid URLs found in the input file or file could not be read."
                print(error_msg)
                logging.error(error_msg)
                return
                
            print("Found {} URLs in the input file.".format(len(urls)))
            logging.info("Found {} URLs in the input file.".format(len(urls)))
        
        # Option 2: Year range specified
        elif year_range:
            start_year, end_year = year_range
            print("Generating URLs for year range: {}-{}".format(start_year, end_year))
            logging.info("Generating URLs for year range: {}-{}".format(start_year, end_year))
            
            # Filter the start_ids and end_ids dictionaries to only include the specified year range
            filtered_start_ids = {year: start_ids.get(year, 1) for year in range(start_year, end_year + 1)}
            if no_auto_discover:
                # Use predefined end IDs
                filtered_end_ids = {year: end_ids.get(year, 5000) for year in range(start_year, end_year + 1)}
                urls = generate_urls_for_year_range(start_year, end_year, filtered_start_ids, filtered_end_ids)
                print("Generated {} URLs for years {}-{}".format(len(urls), start_year, end_year))
                logging.info("Generated {} URLs for years {}-{}".format(len(urls), start_year, end_year))
            else:
                # Auto-discover the end IDs and generate URLs
                urls = auto_discover_urls_for_year_range(start_year, end_year, filtered_start_ids)
                print("Auto-discovered and generated {} URLs for years {}-{}".format(
                    len(urls), start_year, end_year))
                logging.info("Auto-discovered and generated {} URLs for years {}-{}".format(
                    len(urls), start_year, end_year))
        
        # Option 3: Default to all URLs from predefined list or hardcoded list
        elif all_urls:
            # Use the generated list of URLs
            print("Using generated list of URLs.")
            logging.info("Using generated list of URLs.")
            urls = all_urls
            print("Number of URLs: {}".format(len(urls)))
        else:
            # Use the hardcoded list of URLs as fallback
            print("Using hardcoded list of URLs.")
            logging.info("Using hardcoded list of URLs.")
            urls = [
                "https://oag.ca.gov/prop65/60-Day-Notice-2021-02146",
                "https://oag.ca.gov/prop65/60-Day-Notice-2021-02145",
                "https://oag.ca.gov/prop65/60-Day-Notice-2021-02147",
                "https://oag.ca.gov/prop65/60-Day-Notice-2021-02148",
                "https://oag.ca.gov/prop65/60-Day-Notice-2022-02148"
            ]

        # Create queues for multi-threading
        url_queue = Queue.Queue()
        results_queue = Queue.Queue()
        
        # Add URLs to the queue
        for url in urls:
            url_queue.put(url)
        
        # Shared list to track failed URLs (thread-safe with locks)
        failed_urls = []
        
        # Create a counter for tracking progress
        counter = AtomicCounter()
        
        # Determine the number of threads to use
        num_threads = min(threads, len(urls))
        print("Using {} threads for scraping".format(num_threads))
        logging.info("Using {} threads for scraping".format(num_threads))

    # Create and start worker threads
    threads_list = []
    for i in range(num_threads):
        thread = threading.Thread(
            target=url_worker,
            args=(url_queue, results_queue, failed_urls, counter, 3, 5),
            name="Worker-{}".format(i+1)
        )
        thread.daemon = True
        thread.start()
        threads_list.append(thread)
    
    # Track start time for progress estimation
    start_time = time.time()
    total_urls = len(urls)
    
    # Monitor progress while threads are running
    try:
        while any(thread.is_alive() for thread in threads_list):
            # Calculate progress
            processed_count = counter.get()
            if processed_count > 0 and processed_count % 10 == 0:  # Update every 10 URLs
                speed, elapsed, remaining, completion = estimate_progress(start_time, processed_count, total_urls)
                
                print("\n--- PROGRESS UPDATE ---")
                print("Processed {}/{} URLs".format(processed_count, total_urls))
                print("Speed: {:.2f} URLs/minute".format(speed))
                print("Elapsed time: {}".format(elapsed))
                print("Estimated remaining: {}".format(remaining))
                print("Estimated completion: {}".format(completion))
                print("-----------------------\n")
                
            time.sleep(1)  # Check status every second
            
    except KeyboardInterrupt:
        print("\nScraping interrupted by user. Waiting for threads to finish current tasks...")
        # Wait for threads to finish their current tasks
        for thread in threads_list:
            thread.join(10)  # Wait up to 10 seconds for each thread

    # Extract results from the queue
    all_data = []
    while not results_queue.empty():
        all_data.append(results_queue.get())

    # Log summary of results
    logging.info("Scraping completed. Processed {} URLs successfully.".format(len(all_data)))
    
    if failed_urls:
        logging.warning("Failed to process {} URLs:".format(len(failed_urls)))
        for failed_url in failed_urls:
            logging.warning("  - {}".format(failed_url))
        
        # Also write failed URLs to a separate file for easy re-processing
        failed_urls_file = os.path.join(reports_dir, "failed_urls_{}.txt".format(date_stamp))
        with open(failed_urls_file, "w") as f:
            for url in failed_urls:
                f.write(url + "\n")

    # Define output filename with date-only stamp
    if compare_file:
        output_filename = os.path.join(reports_dir, "60-Day-Notices-Updated-{}.xlsx".format(date_stamp))
    else:
        output_filename = os.path.join(reports_dir, "60-Day-Notices-{}.xlsx".format(date_stamp))
    
    # Check if we need to filter data based on comparison
    if comparison_data:
        print("Comparing scraped data with existing data...")
        all_data = compare_and_update_data(all_data, comparison_data)
        if not all_data:
            print("No changes or new data found. Exiting.")
            logging.info("No changes or new data found. Exiting.")
            return
        print("After filtering, we have {} entries with new or changed data".format(len(all_data)))
    
    # Now, extract all possible headers from all data
    all_headers_by_category = {}
    data_categories = ['data', 'flat_civil_complaint_data', 'flat_settlement_data', 
                       'flat_judgment_data', 'flat_corrected_settlement_data']

    # Initialize headers for each category
    for category in data_categories:
        all_headers_by_category[category] = set()

    # Gather all possible headers from all data
    for entry in all_data:
        for category in data_categories:
            if category in entry:
                all_headers_by_category[category].update(entry[category].keys())

    # Remove 'link' from data category since we'll handle it separately
    if 'link' in all_headers_by_category.get('data', []):
        all_headers_by_category['data'].remove('link')

    # Create a new Excel workbook using openpyxl
    workbook = openpyxl.Workbook()
    main_sheet = workbook.active
    if main_sheet:
        main_sheet.title = "Main Data"
    else:
        # If for some reason the active sheet is None, create a new sheet
        main_sheet = workbook.create_sheet("Main Data")
        
    # Write data to the sheet
    write_data_to_sheet_with_all_headers(main_sheet, all_data, all_headers_by_category)

    # Save the Excel file
    workbook.save(output_filename)
    print("Data successfully written to {}".format(output_filename))
    
    # Create MIME attachment for the Excel file
    from email.mime.base import MIMEBase
    from email import encoders
    
    # Create attachment
    with open(output_filename, 'rb') as file:
        attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        attachment.set_payload(file.read())
    
    # Encode file in ASCII characters to send by email    
    encoders.encode_base64(attachment)
    
    # Add header as key/value pair to attachment part
    attachment.add_header(
        'Content-Disposition',
        'attachment; filename="{}"'.format(os.path.basename(output_filename)),
    )
    
    recipients = ("nick@veya.co", "andrew@veya.co", "alexo@veya.co")
    subject = "OAG CA Gov Scraper"
    msg = "The OAG CA Gov Scraper has been completed. The file is attached."
    email_custom(recipients, "OAG_SCRAPER@denile.co", subject, msg, mime_attachments=[attachment])

# Function to load comparison data from an Excel file
def load_comparison_data(excel_file):
    """
    Load comparison data from an Excel file.
    
    Args:
        excel_file: Path to the Excel file
        
    Returns:
        Dictionary mapping URLs to row data
    """
    comparison_data = {}
    
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file)
        if not workbook:
            print("Error: Could not load workbook from file: {}".format(excel_file))
            return comparison_data
            
        # Get the active sheet
        sheet = workbook.active
        if not sheet:
            print("Error: No active sheet found in workbook: {}".format(excel_file))
            return comparison_data
        
        # Get header row from first row
        if sheet.max_row < 1:
            print("Error: Sheet has no data rows")
            return comparison_data
            
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(cell.value)
        
        if not headers:
            print("Error: No headers found in first row")
            return comparison_data
            
        # Look for 'link' column or any column that might contain URLs
        link_index = -1
        possible_url_columns = ['link', 'url', 'Link', 'URL', 'href']
        
        # Try to find an exact column name match first
        for url_col in possible_url_columns:
            if url_col in headers:
                link_index = headers.index(url_col)
                print("Found URL column: {}".format(url_col))
                break
                
        # If no exact match, look for columns containing these words
        if link_index == -1:
            for i, header in enumerate(headers):
                header_str = str(header).lower() if header else ""
                if any(url_term in header_str for url_term in ['link', 'url', 'href']):
                    link_index = i
                    print("Found URL-like column: {}".format(header))
                    break
        
        if link_index == -1:
            print("No 'link' or URL column found in comparison file.")
            return comparison_data
        
        # Read data rows
        max_row = sheet.max_row or 0
        for row_idx in range(2, max_row + 1):
            row_data = {}
            url = None
            
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if not cell:
                    continue
                    
                cell_value = cell.value
                
                # Get URL from the link column
                if col_idx - 1 == link_index:
                    # Handle potential URL format variations
                    if cell_value:
                        cell_value_str = str(cell_value)
                        # Extract URL if it's in markdown format [text](url)
                        md_match = re.search(r'\[.*?\]\((.*?)\)', cell_value_str)
                        if md_match:
                            url = md_match.group(1)
                        # Otherwise use the raw value if it looks like a URL
                        elif 'http' in cell_value_str:
                            url = cell_value_str
                        # If it contains oag.ca.gov/prop65, it's likely a URL fragment
                        elif 'oag.ca.gov/prop65' in cell_value_str:
                            url = cell_value_str
                
                # Store all other values
                row_data[header] = cell_value
            
            # Ensure URL is complete
            if url:
                url_str = str(url)
                if not url_str.startswith('http'):
                    if 'oag.ca.gov/prop65' in url_str:
                        parts = url_str.split('oag.ca.gov/prop65')
                        if len(parts) > 1:
                            url = 'https://oag.ca.gov/prop65' + parts[1].strip()
                    else:
                        # Try to extract the notice ID if it looks like one (e.g., 2021-12345)
                        notice_id_match = re.search(r'(\d{4}[-/]\d{5})', url_str)
                        if notice_id_match:
                            notice_id = notice_id_match.group(1).replace('/', '-')
                            url = "https://oag.ca.gov/prop65/60-Day-Notice-{}".format(notice_id)
            
            # Only add to comparison data if we have a valid URL
            if url and str(url).startswith('http'):
                comparison_data[str(url)] = row_data
                row_data['link'] = url  # Ensure the link is set in the row data
        
        print("Loaded {} records with valid URLs from comparison file.".format(len(comparison_data)))
        return comparison_data
        
    except Exception as e:
        print("Error loading comparison data: {}".format(e))
        logging.error("Error loading comparison data: {}".format(e))
        return {}

# Function to compare and update data
def compare_and_update_data(new_data, comparison_data):
    """
    Compare new data with existing data and update as needed.
    Only returns rows with new data.
    
    Args:
        new_data: List of dictionaries with new data
        comparison_data: Dictionary mapping URLs to existing data
        
    Returns:
        Updated list of data dictionaries containing only new or changed data
    """
    if not comparison_data:
        return new_data
        
    result_data = []
    new_entries_count = 0
    updated_count = 0
    unchanged_count = 0
    
    print("Comparing {} scraped entries against {} existing entries...".format(
        len(new_data), len(comparison_data)))
    
    # Process each entry in the new data
    for entry in new_data:
        url = entry.get('data', {}).get('link')
        
        if not url:
            # This shouldn't happen, but just in case
            print("Warning: Found entry without URL")
            continue
            
        # Convert URL to string for comparison
        url = str(url)
        
        # Check if this URL exists in the comparison data
        if url in comparison_data:
            # Check if any data has changed
            has_changes = False
            
            # What existing data looks like
            existing_entry = comparison_data[url]
            
            # Compare each category of data
            for category in ['data', 'flat_civil_complaint_data', 'flat_settlement_data', 
                           'flat_judgment_data', 'flat_corrected_settlement_data']:
                if category in entry:
                    for key, value in entry[category].items():
                        # Skip the link field for comparison
                        if key == 'link':
                            continue
                            
                        # Get the existing value for this field
                        existing_value = None
                        if category == 'data':
                            # Main data fields are directly in the row
                            existing_value = existing_entry.get(key)
                        else:
                            # For other categories, the fields are prefixed
                            prefixed_key = key  # The key is already prefixed in our data structure
                            existing_value = existing_entry.get(prefixed_key)
                            
                        # Compare the values - if different, mark as changed
                        if value and value != existing_value:
                            has_changes = True
                            break
                            
                if has_changes:
                    break
                    
            if has_changes:
                # Entry has changes, include it with 'Updated' status
                entry['data']['Status'] = 'Updated'
                result_data.append(entry)
                updated_count += 1
                print("Updated entry found: {}".format(url))
            else:
                # No changes, skip this entry
                unchanged_count += 1
        else:
            # URL not in comparison data - this is a new entry
            entry['data']['Status'] = 'New'
            result_data.append(entry)
            new_entries_count += 1
            print("New entry found: {}".format(url))
            
    # Print summary of comparison
    print("\nComparison Summary:")
    print("  - Total entries checked: {}".format(len(new_data)))
    print("  - Entries with updates: {}".format(updated_count))  
    print("  - Unchanged entries (skipped): {}".format(unchanged_count))
    print("  - New entries: {}".format(new_entries_count))
    print("  - Total entries to write: {}".format(len(result_data)))
    
    if not result_data:
        print("\nNo new or updated entries found.")
    
    return result_data

def estimate_progress(start_time, urls_processed, total_urls):
    """
    Estimates scraping speed and time remaining based on progress so far.
    
    Args:
        start_time: Timestamp when scraping started (from time.time())
        urls_processed: Number of URLs successfully processed so far
        total_urls: Total number of URLs to process
    
    Returns:
        Tuple containing:
        - speed: URLs processed per minute
        - elapsed_time_str: Formatted string of elapsed time (HH:MM:SS)
        - remaining_time_str: Formatted string of estimated remaining time (HH:MM:SS)
        - completion_time_str: Formatted string of estimated completion time (HH:MM)
    """
    current_time = time.time()
    elapsed_time = current_time - start_time
    
    # Avoid division by zero
    if urls_processed == 0:
        return 0, "00:00:00", "Unknown", "Unknown"
    
    # Calculate speed (URLs per minute)
    speed = (urls_processed / elapsed_time) * 60
    
    # Calculate remaining time
    if speed > 0:
        remaining_urls = total_urls - urls_processed
        remaining_seconds = (remaining_urls / speed) * 60
        
        # Calculate estimated completion time
        completion_time = current_time + remaining_seconds
        completion_time_str = datetime.datetime.fromtimestamp(completion_time).strftime("%H:%M on %Y-%m-%d")
    else:
        remaining_seconds = float('inf')
        completion_time_str = "Unknown"
    
    # Format elapsed time as HH:MM:SS
    elapsed_time_str = str(datetime.timedelta(seconds=int(elapsed_time)))
    
    # Format remaining time as HH:MM:SS
    if not math.isinf(remaining_seconds):
        remaining_time_str = str(datetime.timedelta(seconds=int(remaining_seconds)))
    else:
        remaining_time_str = "Unknown"
    
    return speed, elapsed_time_str, remaining_time_str, completion_time_str

if __name__ == "__main__":
    main()
